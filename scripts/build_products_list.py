import openpyxl
import json
import re

wb = openpyxl.load_workbook(r'C:\Users\pc\Desktop\Avira Price List official.xlsx')
sheet = wb['Table 1']

# Custom Image mappings for each Avira product category
category_images = {
    'Health & Wellness': 'https://images.unsplash.com/photo-1584308666744-24d5c474f2ae?auto=format&fit=crop&q=80&w=800',
    'Hair Care': 'https://images.unsplash.com/photo-1535585209827-a15fcdbc4c2d?auto=format&fit=crop&q=80&w=800',
    'Personal Care & Skin': 'https://images.unsplash.com/photo-1556228720-195a672e8a03?auto=format&fit=crop&q=80&w=800',
    'Oral Care': 'https://images.unsplash.com/photo-1559591937-e1032b4b455b?auto=format&fit=crop&q=80&w=800',
    'Beverages': 'https://images.unsplash.com/photo-1576092768241-dec231879fc3?auto=format&fit=crop&q=80&w=800',
    'Women Care & Hygiene': 'https://images.unsplash.com/photo-1584308666744-24d5c474f2ae?auto=format&fit=crop&q=80&w=800',
    'Agriculture & Soil Care': 'https://images.unsplash.com/photo-1585314062340-f1a5a7c9328d?auto=format&fit=crop&q=80&w=800'
}

products = []
for r in range(4, sheet.max_row + 1):
    sn = sheet.cell(row=r, column=1).value
    name = sheet.cell(row=r, column=2).value
    vol = sheet.cell(row=r, column=7).value
    mrp_raw = sheet.cell(row=r, column=9).value
    pv_raw = sheet.cell(row=r, column=11).value

    if not sn or not name:
        continue

    name = str(name).strip()
    vol = str(vol or '').strip()
    mrp_str = str(mrp_raw or '').strip()
    pv_str = str(pv_raw or '').strip()

    # Parse MRP (take the discounted selling price if two prices exist, e.g. "2000 1799 RS")
    mrp_nums = re.findall(r'\d+', mrp_str)
    if len(mrp_nums) >= 2:
        mrp = float(mrp_nums[0])
        dp = float(mrp_nums[1])
    elif len(mrp_nums) == 1:
        mrp = float(mrp_nums[0])
        dp = float(mrp_nums[0])
    else:
        mrp = 0.0
        dp = 0.0

    # Parse PV
    pv_nums = re.findall(r'\d+', pv_str)
    pv = float(pv_nums[0]) if pv_nums else 0.0

    # Categorization & HSN Codes
    name_lower = name.lower()
    if any(w in name_lower for w in ['capsule', 'tablet', 'powder', 'drink', 'juice', 'de-addiction', 'jeevan amrut', 'amrut']):
        category = 'Health & Wellness'
        if any(w in name_lower for w in ['capsule', 'tablet', 'drops', 'amrut', 'de-addiction', 'fat loss', 'detox', 'diabetic']):
            hsn = '30049011' # Ayurvedic / Herbal Medicaments
        else:
            hsn = '21069099' # Food / Nutritional Supplements
    elif any(w in name_lower for w in ['shampoo', 'hair oil', 'oil', 'mahendi']):
        category = 'Hair Care'
        if 'shampoo' in name_lower:
            hsn = '33051010'
        elif 'oil' in name_lower:
            hsn = '33059011'
        else:
            hsn = '33059040'
    elif any(w in name_lower for w in ['soap', 'face wash', 'body wash', 'cleanser', 'cream', 'wax', 'pain relief']):
        category = 'Personal Care & Skin'
        if 'soap' in name_lower:
            hsn = '34011110'
        elif 'face wash' in name_lower or 'cleanser' in name_lower or 'cream' in name_lower:
            hsn = '33049910'
        elif 'body wash' in name_lower:
            hsn = '34013011'
        elif 'pain relief' in name_lower:
            hsn = '30049011'
        else:
            hsn = '33079090'
    elif 'toothpaste' in name_lower:
        category = 'Oral Care'
        hsn = '33061010'
    elif 'tea' in name_lower:
        category = 'Beverages'
        hsn = '09024010'
    elif 'napkins' in name_lower or 'pad' in name_lower:
        category = 'Women Care & Hygiene'
        hsn = '96190010'
    elif 'carbonx' in name_lower:
        category = 'Agriculture & Soil Care'
        hsn = '38021000'
    else:
        category = 'General Wellness'
        hsn = '21069099'

    slug = re.sub(r'[^a-z0-9]+', '-', name.lower()).strip('-')

    products.append({
        'id': f"prod_{slug}",
        'sn': sn,
        'name': f"Avira {name}",
        'slug': f"avira-{slug}",
        'net_quantity': vol,
        'mrp': mrp,
        'dp': dp,
        'discount_price': dp,
        'pv': pv,
        'category': category,
        'category_name': category,
        'hsn_code': hsn,
        'image_url': category_images.get(category, 'https://images.unsplash.com/photo-1584308666744-24d5c474f2ae?auto=format&fit=crop&q=80&w=800'),
        'description': f"Official Avira LifeCare premium {name} ({vol}). Crafted with highest grade ingredients for optimal efficacy and results.",
        'stock': 500,
        'stock_quantity': 500,
        'in_stock': True,
        'is_active': True,
        'tag': 'Bestseller' if pv >= 100 else 'Popular'
    })

print(f"Total Products Parsed: {len(products)}")
with open(r'D:\aviracare\avira\scripts\products_catalogue.json', 'w', encoding='utf-8') as f:
    json.dump(products, f, indent=2)

print("Saved to products_catalogue.json")
