import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\pc\Desktop\Avira Price List official.xlsx')
sheet = wb['Table 1']

print("Max row:", sheet.max_row, "Max col:", sheet.max_column)
for r in range(1, 10):
    row_data = []
    for c in range(1, sheet.max_column + 1):
        v = sheet.cell(row=r, column=c).value
        if v is not None:
            row_data.append((c, v))
    print(f"Row {r}: {row_data}")
